from openpyxl import load_workbook
from mysql.connector import connect

excel_path = '/mnt/data/data-samples.xlsx'
wb = load_workbook(excel_path)
sheet = wb.active  # You can also specify a sheet name with wb["Sheet1"]

# === Read Headers ===
headers = [cell.value for cell in sheet[1]]
required_headers = {"equipmentId", "name", "available"}

if not required_headers.issubset(set(headers)):
    raise ValueError("Excel file must have headers: equipmentId, name, available")

# Get column indexes for mapping
header_index = {header: headers.index(header) for header in required_headers}

# === Insert rows into MySQL database ===
db = DatabaseManager()

for row in sheet.iter_rows(min_row=2, values_only=True):
    equipment_id = str(row[header_index["equipmentId"]])
    name = str(row[header_index["name"]])
    available = row[header_index["available"]]

    # If Excel has 'TRUE'/'FALSE' strings instead of boolean
    if isinstance(available, str):
        available = available.strip().lower() in ("true", "yes", "1")

    insert_query = """
    INSERT INTO rental_equipment (equipmentId, name, available)
    VALUES (%s, %s, %s)
    ON DUPLICATE KEY UPDATE name = VALUES(name), available = VALUES(available)
    """
    db.execute_query(insert_query, (equipment_id, name, available))

db.close()
print("✅ All data from Excel has been inserted into the MySQL database.")